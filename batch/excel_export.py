"""
Export crawl results to Excel (JSONL format)
Install: pip install openpyxl pandas
"""

import json
import sys
from pathlib import Path
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def export_jsonl_to_excel(jsonl_file: str, output_file: str = None) -> str:
    """
    Convert JSONL crawl results to Excel file.
    
    Args:
        jsonl_file: Path to JSONL results file
        output_file: Output Excel file path (default: auto-generated)
    
    Returns:
        Path to output Excel file
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: pandas and openpyxl required")
        print("Install with: pip install pandas openpyxl")
        sys.exit(1)
    
    # Read JSONL file
    print(f"📖 Reading {jsonl_file}...")
    
    records = []
    try:
        with open(jsonl_file, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f, 1):
                try:
                    record = json.loads(line)
                    records.append(record)
                except json.JSONDecodeError as e:
                    logger.warning(f"Skipped line {i}: {e}")
                    continue
    except FileNotFoundError:
        print(f"Error: File not found: {jsonl_file}")
        sys.exit(1)
    
    if not records:
        print("No records found in file")
        return None
    
    print(f"✓ Loaded {len(records)} records")
    
    # Convert to DataFrame
    df = pd.DataFrame(records)
    
    # Reorder columns for readability
    column_order = [
        'url', 'companyName', 'email', 'inquiryFormUrl',
        'industry', 'httpStatus', 'crawlStatus',
        'companyNameConfidence', 'emailConfidence', 'industryConfidence',
        'formDetectionMethod', 'lastCrawledAt', 'errorMessage'
    ]
    
    # Only include columns that exist
    available_cols = [col for col in column_order if col in df.columns]
    other_cols = [col for col in df.columns if col not in column_order]
    
    df = df[available_cols + other_cols]
    
    # Generate output filename
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"crawl_results_{timestamp}.xlsx"
    
    # Export to Excel with formatting
    print(f"📝 Exporting to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Results', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set column widths
        col_widths = {
            'A': 35,  # URL
            'B': 25,  # Company Name
            'C': 25,  # Email
            'D': 35,  # Form URL
            'E': 15,  # Industry
            'F': 12,  # HTTP Status
            'G': 12,  # Crawl Status
            'H': 12,  # Company Conf
            'I': 12,  # Email Conf
            'J': 12,  # Industry Conf
            'K': 20,  # Form Method
            'L': 20,  # Last Crawled
            'M': 30,  # Error Message
        }
        
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"✅ Exported to: {output_file}")
    
    # Print summary
    print("\n" + "=" * 70)
    print("EXPORT SUMMARY")
    print("=" * 70)
    print(f"Total records: {len(df)}")
    print(f"File size: {Path(output_file).stat().st_size / (1024*1024):.2f} MB")
    
    # Statistics
    successful = len(df[df['crawlStatus'] == 'success'])
    emails_found = len(df[df['email'].notna()])
    forms_found = len(df[df['inquiryFormUrl'].notna()])
    company_names = len(df[df['companyName'].notna()])
    
    print(f"\nData Statistics:")
    print(f"  Successful crawls: {successful}/{len(df)} ({successful/len(df)*100:.1f}%)")
    print(f"  Emails found: {emails_found} ({emails_found/successful*100:.1f}% of successful)")
    print(f"  Forms found: {forms_found} ({forms_found/successful*100:.1f}% of successful)")
    print(f"  Company names: {company_names} ({company_names/successful*100:.1f}% of successful)")
    
    print("=" * 70 + "\n")
    
    return output_file


def export_csv(jsonl_file: str, output_file: str = None) -> str:
    """Export to CSV instead of Excel."""
    try:
        import pandas as pd
    except ImportError:
        print("Error: pandas required. Install with: pip install pandas")
        sys.exit(1)
    
    print(f"📖 Reading {jsonl_file}...")
    
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    
    if not records:
        print("No records found")
        return None
    
    df = pd.DataFrame(records)
    
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"crawl_results_{timestamp}.csv"
    
    df.to_csv(output_file, index=False, encoding='utf-8-sig')
    print(f"✅ Exported to: {output_file}")
    
    return output_file


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Export crawl results to Excel')
    parser.add_argument('input_file', help='Input JSONL file')
    parser.add_argument('--output', type=str, help='Output file path')
    parser.add_argument('--format', choices=['excel', 'csv'], default='excel',
                       help='Output format (default: excel)')
    
    args = parser.parse_args()
    
    if args.format == 'csv':
        export_csv(args.input_file, args.output)
    else:
        export_jsonl_to_excel(args.input_file, args.output)