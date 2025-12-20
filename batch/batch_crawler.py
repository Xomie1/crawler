"""
Improved Batch Crawler - With automatic Excel export
Processes multiple URLs and exports to Excel automatically
"""

import argparse
import logging
import json
import sys
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import random

try:
    import pandas as pd
except ImportError:
    print("Error: pandas not installed. Run: pip install pandas openpyxl")
    sys.exit(1)

from crawler.engine import CrawlerEngine
from utils.logger import setup_logger

# ==================== LOAD .ENV FILE ====================
import os
from pathlib import Path

def load_dotenv():
    """Load .env file if exists."""
    env_path = Path(__file__).parent / '.env'
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8-sig') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip().strip('"').strip("'")
        print("✓ Loaded .env file")

load_dotenv()
# ========================================================

# Optional Google Sheets export
try:
    from google_sheets_export import GoogleSheetsExporter
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False

# Google Apps Script integration
try:
    from google_apps_script_integration import GoogleAppsScriptIntegration
    GOOGLE_APPS_SCRIPT_AVAILABLE = True
except ImportError:
    GOOGLE_APPS_SCRIPT_AVAILABLE = False

# Setup logging
logger = setup_logger(name="batch_crawler", level=logging.INFO)


class BatchCrawler:
    """Handles batch crawling with automatic Excel export."""
    
    def __init__(
        self,
        timeout: int = 30,
        robots_policy: str = "respect",
        user_agent: str = "CrawlerBot/1.0",
        delay: float = 10.0,
        jitter: float = 2.0,
        use_playwright: bool = True,
        use_enhanced_forms: bool = True,
        max_form_pages: int = 15,
        exclude_patterns: List[str] = None,
        use_ai: bool = False,
        ai_provider: str = 'groq',
        ai_always: bool = False,
        ai_thresholds: Dict[str, float] = None,
        auto_export_excel: bool = True
    ):
        self.timeout = timeout
        self.robots_policy = robots_policy
        self.user_agent = user_agent
        self.delay = delay
        self.jitter = jitter
        self.use_playwright = use_playwright
        self.use_enhanced_forms = use_enhanced_forms
        self.max_form_pages = max_form_pages
        self.results = []
        self.start_time = datetime.now()
        self.exclude_patterns = exclude_patterns if exclude_patterns is not None else []
        self.use_ai = use_ai
        self.ai_provider = ai_provider
        self.ai_always = ai_always
        self.ai_thresholds = ai_thresholds
        self.auto_export_excel = auto_export_excel
        self.jsonl_file = None  # Will be set after saving

    def _get_delay(self) -> float:
        """Calculate delay with random jitter."""
        return self.delay + random.uniform(-self.jitter, self.jitter)
    
    def should_exclude(self, url: str) -> bool:
        return any(pattern in url for pattern in self.exclude_patterns)

    def crawl_urls(self, urls: List[str], company_names: List[str] = None) -> List[Dict]:
        """Crawl multiple URLs with delays."""
        total = len(urls)
        if company_names is None:
            company_names = [None] * total
        
        for i, (url, company_name) in enumerate(zip(urls, company_names), 1):
            try:
                if self.should_exclude(url):
                    logger.info(f"[{i}/{total}] Skipping excluded URL: {url}")
                    self.results.append({
                        'url': url,
                        'email': None,
                        'inquiryFormUrl': None,
                        'companyName': company_name,
                        'industry': None,
                        'httpStatus': None,
                        'robotsAllowed': None,
                        'lastCrawledAt': datetime.utcnow().isoformat(),
                        'crawlStatus': 'skipped',
                        'errorMessage': 'URL matched exclude pattern',
                        'formDetectionMethod': 'skipped'
                    })
                    continue
                
                if i > 1:
                    delay_time = self._get_delay()
                    logger.info(f"[{i}/{total}] Waiting {delay_time:.1f}s before next crawl...")
                    time.sleep(delay_time)
                
                logger.info(f"[{i}/{total}] Crawling: {url}")
                
                crawler = CrawlerEngine(
                    root_url=url,
                    crawl_settings={'timeout': self.timeout},
                    user_agent_policy=self.user_agent,
                    robots_policy=self.robots_policy,
                    use_enhanced_form_detection=self.use_enhanced_forms,
                    max_form_pages=self.max_form_pages,
                    use_ai_extraction=self.use_ai,
                    ai_provider=self.ai_provider,
                    ai_always=self.ai_always,
                    ai_confidence_thresholds=self.ai_thresholds
                )

                crawler.use_playwright = self.use_playwright
                result = crawler.crawl()
                self.results.append(result)
                
                status = result.get('crawlStatus')
                email = result.get('email')
                form = result.get('inquiryFormUrl')
                form_source = result.get('formDetectionMethod', 'unknown')
                
                if status == 'success':
                    form_status = "✓" if form else "✗"
                    logger.info(f"  ✓ Success - Email: {email or 'N/A'}, Form: {form_status} (via: {form_source})")
                else:
                    logger.warning(f"  ✗ Failed - {result.get('errorMessage')}")
                
                crawler.close()
                
            except Exception as e:
                logger.error(f"Error crawling {url}: {e}")
                self.results.append({
                    'url': url,
                    'email': None,
                    'inquiryFormUrl': None,
                    'companyName': company_name,
                    'industry': None,
                    'httpStatus': 0,
                    'robotsAllowed': True,
                    'lastCrawledAt': datetime.utcnow().isoformat(),
                    'crawlStatus': 'error',
                    'errorMessage': str(e),
                    'formDetectionMethod': 'error'
                })
        
        return self.results
    
    def save_results(self, output_file: str = None):
        """Save crawl results to JSONL file."""
        if output_file is None:
            output_file = f"crawl_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl"
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                for result in self.results:
                    f.write(json.dumps(result, ensure_ascii=False) + '\n')
            
            logger.info(f"\n✓ Results saved to: {output_file}")
            self.jsonl_file = output_file
            return output_file
        except Exception as e:
            logger.error(f"Failed to save results: {e}")
            return None
    
    def export_to_excel(self, excel_file: str = None) -> Optional[str]:
        """Export JSONL results to Excel."""
        if not self.jsonl_file:
            logger.error("No JSONL file to export")
            return None
        
        if excel_file is None:
            # Auto-generate filename from JSONL
            base_name = Path(self.jsonl_file).stem
            excel_file = f"{base_name}.xlsx"
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return None
        
        logger.info(f"\n📝 Exporting to Excel: {excel_file}")
        
        try:
            # Convert results to DataFrame
            df = pd.DataFrame(self.results)
            
            # Reorder columns
            column_order = [
                'url', 'companyName', 'email', 'inquiryFormUrl',
                'industry', 'httpStatus', 'crawlStatus',
                'companyNameConfidence', 'emailConfidence', 'industryConfidence',
                'formDetectionMethod', 'lastCrawledAt', 'errorMessage'
            ]
            
            available_cols = [col for col in column_order if col in df.columns]
            other_cols = [col for col in df.columns if col not in column_order]
            
            df = df[available_cols + other_cols]
            
            # Export with formatting
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Results', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Header styling
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format header
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Format data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths
                col_widths = {
                    'A': 35,  # URL
                    'B': 25,  # Company Name
                    'C': 25,  # Email
                    'D': 35,  # Form URL
                    'E': 15,  # Industry
                    'F': 12,  # HTTP Status
                    'G': 12,  # Crawl Status
                    'H': 12,  # Company Conf
                    'I': 12,  # Email Conf
                    'J': 12,  # Industry Conf
                    'K': 20,  # Form Method
                    'L': 20,  # Last Crawled
                    'M': 30,  # Error Message
                }
                
                for col_letter, width in col_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Freeze header
                worksheet.freeze_panes = 'A2'
            
            logger.info(f"✓ Exported to: {excel_file}")
            
            # Print statistics
            self._print_export_stats(df)
            
            return excel_file
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            return None
    
    def _print_export_stats(self, df):
        """Print export statistics."""
        total = len(df)
        successful = len(df[df['crawlStatus'] == 'success'])
        emails_found = len(df[df['email'].notna()])
        forms_found = len(df[df['inquiryFormUrl'].notna()])
        company_names = len(df[df['companyName'].notna()])
        
        print("\n" + "=" * 70)
        print("📊 EXCEL EXPORT SUMMARY")
        print("=" * 70)
        print(f"Total records: {total}")
        print(f"File size: {Path(self.jsonl_file).stat().st_size / (1024*1024):.2f} MB")
        print(f"\nData Statistics:")
        print(f"  Successful crawls: {successful}/{total} ({successful/total*100:.1f}%)")
        if successful > 0:
            print(f"  Emails found: {emails_found}/{successful} ({emails_found/successful*100:.1f}%)")
            print(f"  Forms found: {forms_found}/{successful} ({forms_found/successful*100:.1f}%)")
            print(f"  Company names: {company_names}/{successful} ({company_names/successful*100:.1f}%)")
        print("=" * 70 + "\n")
    
    def generate_summary(self):
        """Generate and print summary statistics."""
        total = len(self.results)
        successful = sum(1 for r in self.results if r.get('crawlStatus') == 'success')
        failed = sum(1 for r in self.results if r.get('crawlStatus') == 'error')
        
        emails_found = sum(1 for r in self.results if r.get('email'))
        forms_found = sum(1 for r in self.results if r.get('inquiryFormUrl'))
        company_names_found = sum(1 for r in self.results if r.get('companyName'))
        
        detected = sum(1 for r in self.results if r.get('formDetectionMethod') == 'enhanced_detector')
        basic = sum(1 for r in self.results if r.get('formDetectionMethod') in ['basic_detector', 'basic_fallback'])
        not_found = sum(1 for r in self.results if r.get('formDetectionMethod') == 'not_found')
        
        elapsed = (datetime.now() - self.start_time).total_seconds()
        
        print("\n" + "=" * 70)
        print("CRAWL RESULTS SUMMARY")
        print("=" * 70)
        print(f"Total URLs: {total}")
        print(f"Successful: {successful} ({successful/total*100 if total > 0 else 0:.1f}%)")
        print(f"Failed: {failed} ({failed/total*100 if total > 0 else 0:.1f}%)")
        print("-" * 70)
        print(f"Emails Found: {emails_found}/{total} ({emails_found/total*100 if total > 0 else 0:.1f}%)")
        print(f"Forms Found: {forms_found}/{total} ({forms_found/total*100 if total > 0 else 0:.1f}%)")
        if detected > 0:
            print(f"  - Enhanced detection: {detected} ({detected/total*100:.1f}%)")
        if basic > 0:
            print(f"  - Basic detection: {basic} ({basic/total*100:.1f}%)")
        if not_found > 0:
            print(f"  - Not found: {not_found} ({not_found/total*100:.1f}%)")
        print(f"Company Names: {company_names_found}/{total} ({company_names_found/total*100 if total > 0 else 0:.1f}%)")
        print("-" * 70)
        print(f"Total Time: {elapsed:.1f}s")
        if total > 0:
            print(f"Avg Time/URL: {elapsed/total:.1f}s")
        print("=" * 70 + "\n")


def load_urls_from_excel(file_path: str, url_column: str = None, limit: int = None) -> tuple:
    """Load URLs from Excel file."""
    try:
        df = pd.read_excel(file_path)
        
        # Auto-detect URL column
        if url_column is None:
            possible_cols = ['トップページURL', 'URL', 'Url', 'url', 'Homepage', 'homepage']
            url_column = None
            for col in possible_cols:
                if col in df.columns:
                    url_column = col
                    break
            
            if url_column is None:
                for col in df.columns:
                    if 'url' in col.lower() or 'homepage' in col.lower():
                        url_column = col
                        break
        
        if url_column is None:
            logger.error(f"Could not find URL column. Available columns: {list(df.columns)}")
            return [], []
        
        logger.info(f"Using URL column: {url_column}")
        
        urls = df[url_column].dropna().astype(str).tolist()
        
        # Try to get company names
        company_names = []
        company_col = None
        for col in ['法人名', 'Company', 'companyName', 'company_name']:
            if col in df.columns:
                company_col = col
                break
        
        if company_col:
            company_names = df[company_col].astype(str).tolist()
            company_names = [name if name != 'nan' else None for name in company_names]
        else:
            company_names = [None] * len(urls)
        
        # Filter and clean URLs
        filtered_urls = []
        filtered_names = []
        for url, name in zip(urls, company_names):
            url_clean = url.strip()
            if url_clean and url_clean != 'nan':
                if not url_clean.startswith(('http://', 'https://')):
                    url_clean = 'https://' + url_clean
                filtered_urls.append(url_clean)
                filtered_names.append(name if name != 'nan' else None)
        
        if limit:
            filtered_urls = filtered_urls[:limit]
            filtered_names = filtered_names[:limit]
        
        logger.info(f"Loaded {len(filtered_urls)} URLs from {file_path}")
        return filtered_urls, filtered_names
        
    except Exception as e:
        logger.error(f"Failed to load URLs from {file_path}: {e}")
        return [], []


def main():
    """Main batch crawler."""
    parser = argparse.ArgumentParser(
        description='Batch crawl multiple websites and export to Excel'
    )
    parser.add_argument('input_file', help='Input Excel or CSV file')
    parser.add_argument('--url-column', type=str, help='Column name with URLs')
    parser.add_argument('--limit', type=int, help='Limit number of URLs to crawl')
    parser.add_argument('--timeout', type=int, default=30, help='Request timeout (default: 30s)')
    parser.add_argument('--delay', type=float, default=10.0, help='Delay between requests (default: 10s)')
    parser.add_argument('--jitter', type=float, default=2.0, help='Random jitter (default: 2s)')
    parser.add_argument('--robots-policy', choices=['respect', 'ignore'], default='respect',
                        help='Robots.txt policy')
    parser.add_argument('--output', type=str, help='Output JSONL file path')
    parser.add_argument('--excel', type=str, help='Output Excel file path (auto if not specified)')
    parser.add_argument('--no-excel', action='store_true', help='Disable auto Excel export')
    parser.add_argument('--no-playwright', action='store_true', help='Disable Playwright')
    parser.add_argument('--enhanced-forms', action='store_true', default=True,
                        help='Use enhanced form detection')
    parser.add_argument('--basic-forms', action='store_true', help='Use basic form detection')
    parser.add_argument('--max-form-pages', type=int, default=15, help='Max form detection pages')
    parser.add_argument('--exclude-patterns', type=str, help='Exclude URL patterns (comma-separated)')
    parser.add_argument('--use-ai', action='store_true', help='Enable AI extraction')
    parser.add_argument('--ai-provider', choices=['groq', 'openai'], default='groq')
    parser.add_argument('--ai-always', action='store_true', help='Always use AI')
    parser.add_argument('--ai-threshold-company', type=float, default=0.7)
    parser.add_argument('--ai-threshold-email', type=float, default=0.5)
    parser.add_argument('--ai-threshold-industry', type=float, default=0.6)
    
    args = parser.parse_args()
    
    # AI setup
    if args.use_ai:
        logger.info(f"\n🤖 AI Extraction: ENABLED")
        logger.info(f"   Provider: {args.ai_provider}")
        logger.info(f"   Mode: {'AI-Only' if args.ai_always else 'Hybrid'}")
        
        from config.ai_config import validate_configuration
        if not validate_configuration(args.ai_provider):
            logger.error(f"❌ {args.ai_provider.upper()}_API_KEY not found")
            sys.exit(1)
        logger.info(f"   ✓ API key validated")
    else:
        logger.info(f"\n🤖 AI Extraction: DISABLED")

    exclude_patterns = []
    if args.exclude_patterns:
        exclude_patterns = [p.strip() for p in args.exclude_patterns.split(',')]

    if not Path(args.input_file).exists():
        logger.error(f"Input file not found: {args.input_file}")
        sys.exit(1)
    
    use_enhanced_forms = args.enhanced_forms and not args.basic_forms
    logger.info(f"\nForm Detection: {'Enhanced' if use_enhanced_forms else 'Basic'}")
    
    urls, company_names = load_urls_from_excel(args.input_file, args.url_column, args.limit)
    
    if not urls:
        logger.error("No URLs to crawl")
        sys.exit(1)
    
    logger.info(f"Starting batch crawl with {len(urls)} URLs...")
    
    ai_thresholds = None
    if args.use_ai:
        ai_thresholds = {
            'company_name': args.ai_threshold_company,
            'email': args.ai_threshold_email,
            'industry': args.ai_threshold_industry
        }

    # Run crawler
    crawler = BatchCrawler(
        timeout=args.timeout,
        robots_policy=args.robots_policy,
        delay=args.delay,
        jitter=args.jitter,
        use_playwright=not args.no_playwright,
        use_enhanced_forms=use_enhanced_forms,
        max_form_pages=args.max_form_pages,
        exclude_patterns=exclude_patterns,
        use_ai=args.use_ai,
        ai_provider=args.ai_provider,
        ai_always=args.ai_always,
        ai_thresholds=ai_thresholds,
        auto_export_excel=not args.no_excel
    )
    
    results = crawler.crawl_urls(urls, company_names)
    
    # Save JSONL
    output_file = crawler.save_results(args.output)
    
    # Print summary
    crawler.generate_summary()
    
    # Export to Excel (automatic unless disabled)
    if crawler.auto_export_excel and output_file:
        excel_file = crawler.export_to_excel(args.excel)
        if excel_file:
            print(f"\n📂 All files saved:")
            print(f"  JSONL: {output_file}")
            print(f"  Excel: {excel_file}")
    
    # Print sample results
    print("\nSample Results (first 3):")
    for result in results[:3]:
        print(f"\n  URL: {result['url']}")
        print(f"    Status: {result['crawlStatus']}")
        print(f"    Company: {result.get('companyName') or 'N/A'}")
        print(f"    Email: {result.get('email') or 'N/A'}")
        print(f"    Form: {result.get('inquiryFormUrl') or 'N/A'}")


if __name__ == "__main__":
    main()